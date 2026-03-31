"""
AutoTax — 엑셀 생성기
plan.md §9.1 홈택스 간이지급명세서 11컬럼 엑셀 양식 + 기안용 양식
"""
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from db.repository import Repository
from core.crypto import CryptoManager
import csv


# ─────────────────────────────────────────────
# 홈택스 간이지급명세서 엑셀 (plan.md §9.1)
# ─────────────────────────────────────────────

HOMETAX_HEADERS = [
    '일련번호', '귀속연도', '귀속월', '업종코드', '소득자성명',
    '주민등록번호', '내외국인', '지급액', '세율', '소득세', '지방소득세'
]


def generate_hometax_excel(
    repo: Repository,
    period: str,
    output_dir: str = None,
) -> str:
    """
    홈택스 간이지급명세서(사업소득) 엑셀 파일 생성.
    plan.md §9.1 — 정확히 11컬럼(A~K) 양식.

    Args:
        repo: Repository 인스턴스
        period: 'YYYY-MM'
        output_dir: 저장 디렉토리 (None이면 프로젝트 루트)

    Returns:
        생성된 파일 경로
    """
    settlements = repo.get_settlements_by_period(period)
    if not settlements:
        raise ValueError(f'{period} 기간에 정산 데이터가 없습니다.')

    year, month = period.split('-')

    if output_dir is None:
        output_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '간이지급명세서'

    # 헤더 스타일
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 헤더 행
    for col, header in enumerate(HOMETAX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터 행
    for idx, s in enumerate(settlements, 1):
        row = idx + 1

        # 주민번호
        rid = s.get('resident_id', '')

        # 주민번호 하이픈 형식 (홈택스 요구)
        if len(rid) == 13 and '-' not in rid:
            rid_formatted = f'{rid[:6]}-{rid[6:]}'
        else:
            rid_formatted = rid

        values = [
            idx,                    # A: 일련번호
            year,                   # B: 귀속연도
            month,                  # C: 귀속월
            s['industry_code'],     # D: 업종코드
            s.get('name', ''),      # E: 소득자성명
            rid_formatted,          # F: 주민등록번호
            s['is_foreigner'],      # G: 내외국인
            s['total_payment'],     # H: 지급액
            s['tax_rate'],          # I: 세율
            s['final_income_tax'],  # J: 소득세
            s['final_local_tax'],   # K: 지방소득세
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            # 숫자 우측 정렬
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal='center')

    # 열 너비 자동 조정
    col_widths = [8, 10, 8, 10, 15, 18, 10, 15, 8, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 파일명: 1000건 초과 시 분할 (plan.md §6.2)
    filename = f'간이지급명세서_{period}.xlsx'
    filepath = os.path.join(output_dir, filename)

    wb.save(filepath)
    wb.save(filepath)
    return filepath


def generate_hometax_csv(
    repo: Repository,
    period: str,
    output_dir: str = None,
) -> str:
    """
    홈택스 간이지급명세서(사업소득) CSV 파일 생성.
    사용자 요청에 따라 변환파일 제출용으로 CSV를 생성합니다.
    """
    settlements = repo.get_settlements_by_period(period)
    if not settlements:
        raise ValueError(f'{period} 기간에 정산 데이터가 없습니다.')

    year, month = period.split('-')

    if output_dir is None:
        output_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(output_dir, exist_ok=True)

    filename = f'간이지급명세서_{period}.csv'
    filepath = os.path.join(output_dir, filename)

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(HOMETAX_HEADERS)

        for idx, s in enumerate(settlements, 1):
            rid = s.get('resident_id', '')

            if len(rid) == 13 and '-' not in rid:
                rid_formatted = f'{rid[:6]}-{rid[6:]}'
            else:
                rid_formatted = rid

            values = [
                idx,
                year,
                month,
                s['industry_code'],
                s.get('name', ''),
                rid_formatted,
                s['is_foreigner'],
                s['total_payment'],
                s['tax_rate'],
                s['final_income_tax'],
                s['final_local_tax'],
            ]
            writer.writerow(values)

    return filepath


# ─────────────────────────────────────────────
# 기안용 강사료 지급내역 엑셀 (prototype.html 양식)
# ─────────────────────────────────────────────

CUSTOM_HEADERS = [
    '연번', '프로그램', '강사', '산출근거(1회강사료)',
    '산출근거(강의횟수)', '강사료', '세액(소득세)',
    '세액(주민세)', '세액(합계)', '실지급액', '계좌번호'
]


def generate_custom_excel(
    repo: Repository,
    period: str,
    category_filter: str = None,
    output_dir: str = None,
) -> str:
    """
    기안용 강사료 지급내역 엑셀 생성.
    plan.md §10.5 + prototype.html generateCustomLectureExcel() 이식.

    Args:
        repo: Repository
        period: 'YYYY-MM'
        category_filter: 과목구분 필터 (None이면 전체)
        output_dir: 저장 디렉토리

    Returns:
        생성된 파일 경로
    """
    from core.tax_calculator import calculate_taxes, get_tax_rate

    lectures = repo.get_lectures_by_period(period)
    if not lectures:
        raise ValueError(f'{period} 기간에 강의 데이터가 없습니다.')

    # 필터 적용
    if category_filter:
        lectures = [l for l in lectures
                    if l.get('program_category', '') == category_filter]

    year, month = period.split('-')
    cat_label = category_filter or '전체과목'

    if output_dir is None:
        output_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '기안용 강의내역'

    # 타이틀 행
    title_text = f'{year}년  {int(month)}월 {cat_label} 강사료 지급내역'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 빈 행
    # 단위 행
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
    unit_cell = ws.cell(row=3, column=1, value='(단위 : 원)')
    unit_cell.alignment = Alignment(horizontal='right')
    unit_cell.font = Font(size=9)

    # 헤더
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 헤더 1행 (Row 4)
    h1_values = {
        1: '연번', 2: '프로그램', 3: '강사', 4: '산출근거', 
        6: '강사료', 7: '세액', 10: '실지급액', 11: '계좌번호'
    }
    # 헤더 2행 (Row 5)
    h2_values = {
        4: '1회강사료', 5: '강의횟수', 7: '소득세', 8: '주민세', 9: '합계'
    }

    for r in [4, 5]:
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if r == 4 and c in h1_values:
                cell.value = h1_values[c]
            elif r == 5 and c in h2_values:
                cell.value = h2_values[c]

    # 셀 병합 (헤더)
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)   # 연번
    ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)   # 프로그램
    ws.merge_cells(start_row=4, start_column=3, end_row=5, end_column=3)   # 강사
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)   # 산출근거
    ws.merge_cells(start_row=4, start_column=6, end_row=5, end_column=6)   # 강사료
    ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=9)   # 세액
    ws.merge_cells(start_row=4, start_column=10, end_row=5, end_column=10) # 실지급액
    ws.merge_cells(start_row=4, start_column=11, end_row=5, end_column=11) # 계좌번호

    # 데이터
    data_row = 6
    for idx, lec in enumerate(lectures, 1):
        total = lec['payment_amount']
        rate = get_tax_rate(lec['industry_code'])
        taxes = calculate_taxes(total, rate)
        total_tax = taxes['income_tax'] + taxes['local_tax']

        # 계좌정보 (줄바꿈 처리)
        bank = lec.get('bank_name', '') or ''
        account = lec.get('account_number', '') or ''
        bank_info = f'{bank}\n{account}'.strip() if bank or account else ''

        values = [
            idx,
            lec.get('program_name', ''),
            lec.get('instructor_name', ''),
            lec['fee_per_session'],
            lec['session_count'],
            total,
            taxes['income_tax'],
            taxes['local_tax'],
            total_tax,
            taxes['net_payment'],
            bank_info,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.border = thin_border
            if isinstance(val, (int, float)) and col > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'
            elif col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 계좌번호 등 텍스트
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        data_row += 1

    # 열 너비
    col_widths = [6, 18, 10, 14, 10, 14, 12, 12, 12, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    prefix = f"{category_filter} " if category_filter else ""
    filename = f'{prefix}강사료 지급내역_{period}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ─────────────────────────────────────────────
# 연간 거주자 사업소득 지급명세서 (prototype.html 양식)
# ─────────────────────────────────────────────

def generate_annual_excel(
    repo: Repository,
    year: str,
    months: list[str] = None,
    output_dir: str = None,
) -> str:
    """
    연간 거주자 사업소득 지급명세서 엑셀 생성.

    Args:
        repo: Repository
        year: '2026'
        months: 선택 월 리스트 (None이면 전체)
        output_dir: 저장 디렉토리

    Returns:
        생성된 파일 경로
    """
    data = repo.get_annual_summary(year, months)
    if not data:
        raise ValueError(f'{year}년에 정산 데이터가 없습니다.')

    if output_dir is None:
        output_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '연간거주자사업소득'

    headers = [
        '주민번호', '강사명', '업종코드',
        '연간 총지급액', '연간 총소득세', '연간 총지방소득세', '연간 총실지급액'
    ]

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for idx, row_data in enumerate(data, 2):
        # 주민번호
        rid = row_data.get('resident_id', '')

        if len(rid) == 13 and '-' not in rid:
            rid = f'{rid[:6]}-{rid[6:]}'

        values = [
            rid,
            row_data.get('name', ''),
            row_data.get('industry_code', ''),
            row_data.get('annual_total', 0),
            row_data.get('annual_income_tax', 0),
            row_data.get('annual_local_tax', 0),
            row_data.get('annual_net_payment', 0),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'

    col_widths = [18, 12, 10, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    month_label = '_'.join(months) if months else '전체'
    filename = f'연간거주자사업소득지급명세서_{year}_{month_label}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath

# ─────────────────────────────────────────────
# 강사 일괄 등록용 (Instructor Batch Register)
# ─────────────────────────────────────────────

INSTRUCTOR_TEMPLATE_HEADERS = [
    '강사명*', '주민번호*', '업종코드*', '연락처', '이메일', '주소',
    '은행', '계좌번호', '과목구분*', '프로그램명*', '회당강사료*', '비고'
]

def generate_instructor_template(output_path: str) -> str:
    """강사 등록용 엑셀 양식 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = '강사등록양식'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    for col, h in enumerate(INSTRUCTOR_TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 예시 데이터 (한 행)
    ws.append(['홍길동', '800101-1234567', '940909', '010-1234-5678', 'test@test.com', '서울시...', '신한은행', '110-123-456789', '인문학', '시낭송', 50000, '비고'])
    
    col_widths = [12, 18, 10, 15, 20, 30, 12, 20, 15, 20, 15, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        
    wb.save(output_path)
    return output_path

def parse_instructor_excel(file_path: str) -> list[dict]:
    """강사 등록 엑셀 파일 파싱"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    instructors = []
    # 2번 행부터 데이터
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name: continue
        
        data = {
            'name': str(name).strip(),
            'resident_id': str(ws.cell(row=row, column=2).value or '').strip(),
            'industry_code': str(ws.cell(row=row, column=3).value or '940909').strip(),
            'phone': str(ws.cell(row=row, column=4).value or '').strip(),
            'email': str(ws.cell(row=row, column=5).value or '').strip(),
            'address': str(ws.cell(row=row, column=6).value or '').strip(),
            'bank_name': str(ws.cell(row=row, column=7).value or '').strip(),
            'account_number': str(ws.cell(row=row, column=8).value or '').strip(),
            'category': str(ws.cell(row=row, column=9).value or '').strip(),
            'program_name': str(ws.cell(row=row, column=10).value or '').strip(),
            'fee_per_session': int(ws.cell(row=row, column=11).value or 0),
            'memo': str(ws.cell(row=row, column=12).value or '').strip(),
        }
        instructors.append(data)
        
    return instructors
def generate_sample_instructor_excel(output_path: str) -> str:
    """강사 등록용 샘플 엑셀 파일 생성 (더미 데이터 포함)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "강사등록_양식"
    
    headers = [
        "강사명", "주민등록번호", "업종코드", "연락처", "이메일", "주소", 
        "은행명", "계좌번호", "과목구분", "프로그램명", "회당 강사료", "메모"
    ]
    
    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # 샘플 데이터 1
    sample_data = [
        ["홍길동", "800101-1234567", "940909", "010-1234-5678", "hong@example.com", 
         "서울시 강남구 테헤란로", "국민은행", "123-45-67890", "인문학", "조선 역사 산책", 50000, "본관 수업"],
        ["성춘향", "850505-2345678", "940909", "010-9876-5432", "sung@example.com", 
         "서울시 서초구 반포동", "신한은행", "987-65-43210", "음악", "전통 민요 교실", 45000, "금요일 오후"]
    ]
    
    for r_idx, row_data in enumerate(sample_data, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if c_idx == 11: # 강사료 숫자 포맷
                cell.number_format = '#,##0'
                
    # 컬럼 너비 조정
    from openpyxl.utils import get_column_letter
    col_widths = [12, 18, 10, 15, 25, 35, 12, 20, 15, 20, 15, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        
    wb.save(output_path)
    return output_path
